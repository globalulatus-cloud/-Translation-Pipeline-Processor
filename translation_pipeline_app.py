import io
import re
import zipfile
from copy import copy

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Alignment, Side
from openpyxl.utils import get_column_letter


st.set_page_config(
    page_title="Translation Pipeline Processor",
    layout="wide",
    page_icon="🌐",
)

st.markdown(
    """
<style>
    .main-title { font-size: 2rem; font-weight: 700; color: #1a1a2e; }
    .err-badge {
        display:inline-block; background:#fee2e2; color:#991b1b;
        border-radius:20px; padding:3px 12px; font-size:0.8rem; margin:2px;
    }
</style>
""",
    unsafe_allow_html=True,
)

# Canonical internal column names
PRIMARY_PREFIX = "Audit_2"
FALLBACK_PREFIX = "Audit_1"

PRIMARY_SCORE = f"{PRIMARY_PREFIX}_score"
PRIMARY_RISK = f"{PRIMARY_PREFIX}_riskLevel"
PRIMARY_ISSUES = f"{PRIMARY_PREFIX}_clinicalIssues"
PRIMARY_TRANSLATION = f"{PRIMARY_PREFIX}_suggestedTranslation"

FALLBACK_SCORE = f"{FALLBACK_PREFIX}_score"
FALLBACK_RISK = f"{FALLBACK_PREFIX}_riskLevel"
FALLBACK_ISSUES = f"{FALLBACK_PREFIX}_clinicalIssues"
FALLBACK_TRANSLATION = f"{FALLBACK_PREFIX}_suggestedTranslation"

REQUIRED_COLS = {
    PRIMARY_SCORE,
    PRIMARY_RISK,
    PRIMARY_ISSUES,
    PRIMARY_TRANSLATION,
    FALLBACK_SCORE,
    FALLBACK_RISK,
    FALLBACK_ISSUES,
    FALLBACK_TRANSLATION,
}

MODEL_COLS = REQUIRED_COLS.copy()

SOURCE_NAME_HINTS = [
    "source", "src", "original", "input", "text", "segment", "utterance", "sentence"
]

CJK_RANGES = [
    ("\u3040", "\u30FF"),   # Hiragana + Katakana
    ("\u3400", "\u4DBF"),   # CJK Extension A
    ("\u4E00", "\u9FFF"),   # CJK Unified Ideographs
    ("\uAC00", "\uD7AF"),   # Hangul syllables
    ("\uF900", "\uFAFF"),   # CJK Compatibility Ideographs
    ("\u0E00", "\u0E7F"),   # Thai
    ("\U00020000", "\U0002A6DF"),  # CJK Extension B
    ("\U0002A700", "\U0002CEAF"),   # CJK Extensions C/D/E
]


def _is_cjk_char(ch: str) -> bool:
    return any(lo <= ch <= hi for lo, hi in CJK_RANGES)


def count_units(text) -> int:
    """
    Count words for space-delimited scripts, and count characters for CJK/Thai-heavy text.
    """
    if pd.isna(text):
        return 0

    s = str(text).strip()
    if not s:
        return 0

    non_ws_chars = [ch for ch in s if not ch.isspace()]
    if not non_ws_chars:
        return 0

    cjk_count = sum(1 for ch in non_ws_chars if _is_cjk_char(ch))
    ratio = cjk_count / len(non_ws_chars)

    if ratio >= 0.20:
        return len(non_ws_chars)
    return len([tok for tok in re.split(r"\s+", s) if tok])


def normalize_col_name(name: str) -> str:
    """
    Normalize column names so variants like:
    - Audit_Score_1
    - Audit Score 1
    - Audit_Score-1
    - RiskLevel-1
    - Suggested_translation_1
    all map to a consistent comparison format.
    """
    s = str(name).strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def standardize_audit_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename incoming Audit columns into a canonical internal schema.

    Expected canonical schema:
    - Audit_1_score
    - Audit_1_riskLevel
    - Audit_1_clinicalIssues
    - Audit_1_suggestedTranslation
    - Audit_2_score
    - Audit_2_riskLevel
    - Audit_2_clinicalIssues
    - Audit_2_suggestedTranslation
    """
    aliases = {
        PRIMARY_SCORE: {
            "audit_score_2", "audit2_score", "audit_2_score", "auditscore2", "audit_score2"
        },
        PRIMARY_RISK: {
            "risklevel_2", "risk_level_2", "audit_2_risklevel", "audit2_risklevel", "audit_risklevel_2"
        },
        PRIMARY_ISSUES: {
            "audit_issues_2", "auditissues2", "audit_2_clinicalissues", "audit2_clinicalissues"
        },
        PRIMARY_TRANSLATION: {
            "suggested_translation_2", "suggestedtranslation2", "audit_2_suggestedtranslation", "audit2_suggestedtranslation"
        },
        FALLBACK_SCORE: {
            "audit_score_1", "audit1_score", "audit_1_score", "auditscore1", "audit_score1"
        },
        FALLBACK_RISK: {
            "risklevel_1", "risk_level_1", "audit_1_risklevel", "audit1_risklevel", "audit_risklevel_1"
        },
        FALLBACK_ISSUES: {
            "audit_issues_1", "auditissues1", "audit_1_clinicalissues", "audit1_clinicalissues"
        },
        FALLBACK_TRANSLATION: {
            "suggested_translation_1", "suggestedtranslation1", "audit_1_suggestedtranslation", "audit1_suggestedtranslation"
        },
    }

    norm_to_actual = {}
    for c in df.columns:
        norm_to_actual[normalize_col_name(c)] = c

    rename_map = {}
    for canonical, norm_aliases in aliases.items():
        for norm_alias in norm_aliases:
            if norm_alias in norm_to_actual:
                rename_map[norm_to_actual[norm_alias]] = canonical
                break

    return df.rename(columns=rename_map)


def detect_source_column(columns) -> str | None:
    """
    Auto-detect the source text column from common names, excluding model-specific fields.
    """
    cols = list(columns)
    lowered = {c.lower(): c for c in cols}

    exact_candidates = [
        "source",
        "source text",
        "source_text",
        "src",
        "src text",
        "original",
        "original text",
        "input",
        "input text",
        "text",
        "segment",
        "utterance",
        "sentence",
    ]
    for candidate in exact_candidates:
        if candidate in lowered:
            return lowered[candidate]

    for c in cols:
        cl = c.lower()
        if any(hint in cl for hint in SOURCE_NAME_HINTS):
            if c not in MODEL_COLS:
                return c

    for c in cols:
        cl = c.lower()
        if c in MODEL_COLS:
            continue
        if "glossary" in cl:
            continue
        if cl in {"source_file", "source text", "source_text"}:
            continue
        return c

    return None


def copy_worksheet_into_workbook(source_ws, target_wb, title="Glossary"):
    """Clone a worksheet into another workbook, preserving content and common formatting."""
    if source_ws is None:
        return None

    if title in target_wb.sheetnames:
        del target_wb[title]

    target_ws = target_wb.create_sheet(title=title)

    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws[source_cell.coordinate]
            target_cell.value = source_cell.value

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)

    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width
        target_ws.column_dimensions[col_letter].hidden = dim.hidden
        target_ws.column_dimensions[col_letter].bestFit = dim.bestFit
        target_ws.column_dimensions[col_letter].outline_level = dim.outline_level
        target_ws.column_dimensions[col_letter].collapsed = dim.collapsed

    for row_idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx].height = dim.height
        target_ws.row_dimensions[row_idx].hidden = dim.hidden
        target_ws.row_dimensions[row_idx].outline_level = dim.outline_level
        target_ws.row_dimensions[row_idx].collapsed = dim.collapsed

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    if source_ws.freeze_panes:
        target_ws.freeze_panes = source_ws.freeze_panes

    try:
        target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    except Exception:
        pass

    return target_ws


def run_pipeline(df: pd.DataFrame, primary_min: int, fallback_min: int, filename: str, source_col: str):
    stats = {"filename": filename, "input_rows": len(df), "source_col": source_col}

    df = df[df[source_col].notna()].copy()

    df[PRIMARY_SCORE] = pd.to_numeric(df[PRIMARY_SCORE], errors="coerce")
    df[FALLBACK_SCORE] = pd.to_numeric(df[FALLBACK_SCORE], errors="coerce")

    primary_invalid = df[PRIMARY_SCORE].isna() | (df[PRIMARY_SCORE] < primary_min)
    fallback_zero_or_missing = df[FALLBACK_SCORE].isna() | (df[FALLBACK_SCORE] == 0)
    df = df[~(primary_invalid & fallback_zero_or_missing)].copy()
    stats["after_step1"] = len(df)

    mask0 = df[FALLBACK_SCORE] == 0
    for target_col, source_col_name in [
        (FALLBACK_SCORE, PRIMARY_SCORE),
        (FALLBACK_RISK, PRIMARY_RISK),
        (FALLBACK_ISSUES, PRIMARY_ISSUES),
        (FALLBACK_TRANSLATION, PRIMARY_TRANSLATION),
    ]:
        # FIX: Some pipeline files leave the Audit_1_* text columns completely
        # blank, so pandas infers them as float64 (all-NaN). Newer pandas
        # (2.x / 3.x) will then raise:
        #   "Invalid value [...] for dtype 'float64'"
        # when we try to write Audit_2 strings into that column, because it
        # no longer silently upcasts a numeric column to object dtype on
        # assignment. Force the target column to a compatible dtype first.
        if df[target_col].dtype != df[source_col_name].dtype:
            df[target_col] = df[target_col].astype(object)
        df.loc[mask0, target_col] = df.loc[mask0, source_col_name]

    stats["fallback_rows_updated"] = int(mask0.sum())

    df = df[df[FALLBACK_SCORE] >= fallback_min].copy()
    stats["final_segments"] = len(df)

    df = df.rename(
        columns={
            source_col: "Source Text",
            FALLBACK_ISSUES: "Clinical Issues",
            FALLBACK_TRANSLATION: "Suggested Translation",
        }
    )

    cols_to_drop = [
        PRIMARY_SCORE,
        PRIMARY_RISK,
        PRIMARY_ISSUES,
        PRIMARY_TRANSLATION,
        FALLBACK_SCORE,
        FALLBACK_RISK,
    ]
    df = df.drop(columns=[c for c in cols_to_drop if c in df.columns])

    for c in [col for col in df.columns if "score" in col.lower() or "risk" in col.lower()]:
        if c not in {"Source File"}:
            df = df.drop(columns=[c])

    if "Clinical Issues" not in df.columns:
        df["Clinical Issues"] = ""
    if "Suggested Translation" not in df.columns:
        df["Suggested Translation"] = ""

    source_text_col = "Source Text"
    if source_text_col in df.columns:
        src_pos = df.columns.get_loc(source_text_col)
        df.insert(src_pos + 1, "Source Count", df[source_text_col].apply(count_units))
        stats["total_source_units"] = int(df["Source Count"].sum())
    else:
        df.insert(1, "Source Count", 0)
        stats["total_source_units"] = 0

    df.insert(0, "Source_File", filename)

    stats["count_col"] = "Source Count"
    return df, stats


def style_excel(df: pd.DataFrame, sheet_name: str = "Linguist_Segments", glossary_ws=None) -> bytes:
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])

    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb[sheet_name[:31]]

    hdr_fill = PatternFill("solid", fgColor="1a3a6b")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"

    if glossary_ws is not None:
        copy_worksheet_into_workbook(glossary_ws, wb, title="Glossary")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_zip(results: list) -> bytes:
    buf = io.BytesIO()

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        frames = []
        first_glossary = None

        for df, stats, glossary_ws in results:
            fname = stats["filename"].replace(".xlsx", "")
            zf.writestr(
                f"{fname}_processed.xlsx",
                style_excel(df, sheet_name=fname[:31], glossary_ws=glossary_ws),
            )
            frames.append(df)
            if first_glossary is None and glossary_ws is not None:
                first_glossary = glossary_ws

        if frames:
            merged = pd.concat(frames, ignore_index=True)
            zf.writestr(
                "ALL_FILES_MERGED.xlsx",
                style_excel(merged, "All_Segments", glossary_ws=first_glossary),
            )
            zf.writestr("ALL_FILES_MERGED.csv", merged.to_csv(index=False).encode("utf-8-sig"))

    return buf.getvalue()


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Pipeline Settings")
    primary_min = st.slider("Min primary score", 1, 10, 1)
    fallback_min = st.slider("Min fallback score", 1, 10, 3)
    st.caption("Rows where both scores are below threshold are dropped.")

    st.divider()
    st.markdown("## 🔎 Source Column")
    source_col_override = st.text_input(
        "Source column name (leave blank to auto-detect)",
        value="",
        placeholder="e.g. Source Text, Original, Input Text",
        help="If blank, the app tries to detect the source column automatically.",
    ).strip()

    st.divider()
    st.markdown(
        """
**Pipeline steps**
1. Drop rows where both scores are below threshold
2. Merge Audit 2 into Audit 1 fields when Audit 1 score is 0
3. Remove model-specific columns
4. Keep Audit 1 score at or above the selected minimum
5. Count source text as words or characters automatically
6. Tag source filename
"""
    )

# ── Main UI ───────────────────────────────────────────────────────────────────
st.markdown(
    '<div class="main-title">🌐 Translation Pipeline Processor</div>',
    unsafe_allow_html=True,
)
st.markdown("Upload one or many pipeline Excel files. Each file is processed separately and also merged.")
st.divider()

st.markdown("### 📂 Upload Pipeline Files")
uploaded_files = st.file_uploader(
    "Drop one or multiple .xlsx files",
    type=["xlsx"],
    accept_multiple_files=True,
    help="All files must have the same column structure as the pipeline template.",
)

if not uploaded_files:
    st.info("Upload one or more Excel pipeline files to begin.")
    with st.expander("What does this app do?"):
        st.markdown(
            """
**Supports bulk upload** - drop as many `.xlsx` pipeline files as needed.

Each file goes through the same pipeline:
1. Filter out rows where both scores are 0 or blank
2. Merge Audit 2 data into Audit 1 columns where Audit 1 score = 0
3. Remove model-specific columns
4. Keep only rows that meet the fallback threshold
5. Count source text automatically as words or characters
6. Tag each row with its source filename

**Outputs:**
- Per-file download, Excel and CSV
- Combined merged Excel and CSV
- All files zipped into one download
"""
        )
    st.stop()

results, errors = [], []
progress = st.progress(0, text="Processing files...")

for i, f in enumerate(uploaded_files):
    try:
        file_bytes = f.getvalue()
        xls = pd.ExcelFile(io.BytesIO(file_bytes))

        main_sheet = xls.sheet_names[0]
        df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=main_sheet, header=0)
        df_raw = standardize_audit_columns(df_raw)

        missing = REQUIRED_COLS - set(df_raw.columns)
        if missing:
            errors.append((f.name, f"Missing columns: {', '.join(sorted(missing))}"))
            progress.progress((i + 1) / len(uploaded_files), text=f"Processed: {f.name}")
            continue

        source_col = source_col_override or detect_source_column(df_raw.columns.tolist())
        if source_col is None or source_col not in df_raw.columns:
            errors.append((f.name, "Missing source column. Set it manually in the sidebar or rename it to a recognizable source field."))
            progress.progress((i + 1) / len(uploaded_files), text=f"Processed: {f.name}")
            continue

        source_wb = load_workbook(io.BytesIO(file_bytes))
        glossary_ws = source_wb["Glossary"] if "Glossary" in source_wb.sheetnames else None

        df_out, stats = run_pipeline(df_raw, primary_min, fallback_min, f.name, source_col)
        results.append((df_out, stats, glossary_ws))
    except Exception as e:
        errors.append((f.name, str(e)))

    progress.progress((i + 1) / len(uploaded_files), text=f"Processed: {f.name}")

progress.empty()

if errors:
    st.error(f"⚠️ {len(errors)} file(s) skipped due to errors:")
    for fname, msg in errors:
        st.markdown(f'<span class="err-badge">❌ {fname}</span> - {msg}', unsafe_allow_html=True)

if not results:
    st.warning("No files were processed successfully. Please check your uploads.")
    st.stop()

# ── Summary ──────────────────────────────────────────────────────────────────
st.markdown("### 📊 Overall Summary")
total_input = sum(s["input_rows"] for _, s, _ in results)
total_final = sum(s["final_segments"] for _, s, _ in results)
total_updated = sum(s["fallback_rows_updated"] for _, s, _ in results)
total_units = sum(s["total_source_units"] for _, s, _ in results)

c1, c2, c3, c4 = st.columns(4)
c1.metric("📁 Files Processed", len(results))
c2.metric("📥 Total Input Rows", total_input)
c3.metric("🔄 Fallback Rows Updated", total_updated)
c4.metric("✂️ Final Segments", total_final)
st.caption(f"Total source units: {total_units:,}")

all_dfs = pd.concat([df for df, _, _ in results], ignore_index=True)

st.divider()
st.markdown("### 📑 Results by File")
tab_labels = [s["filename"][:28] for _, s, _ in results]
tabs = st.tabs(["🗂️ ALL MERGED"] + tab_labels)

with tabs[0]:
    st.markdown(f"**{len(all_dfs)} segments across {len(results)} file(s)**")
    sf = st.text_input("Search Source Text", key="merged_search")
    df_view = all_dfs.copy()
    if sf and "Source Text" in df_view.columns:
        df_view = df_view[df_view["Source Text"].astype(str).str.contains(sf, na=False)]

    st.dataframe(df_view, use_container_width=True, height=420)
    st.caption(f"Showing {len(df_view)} of {len(all_dfs)} segments")

for idx, (tab, (df, stats, glossary_ws)) in enumerate(zip(tabs[1:], results)):
    with tab:
        m1, m2, m3 = st.columns(3)
        m1.metric("Input Rows", stats["input_rows"])
        m2.metric("Fallback Rows Updated", stats["fallback_rows_updated"])
        m3.metric("Final Segments", stats["final_segments"])

        st.caption(f"Source column used: `{stats['source_col']}`")

        search_txt = st.text_input("Search Source Text", key=f"s_{idx}")
        dp = df.copy()
        if search_txt and "Source Text" in dp.columns:
            dp = dp[dp["Source Text"].astype(str).str.contains(search_txt, na=False)]

        st.dataframe(dp, use_container_width=True, height=380)
        st.caption(f"Showing {len(dp)} of {stats['final_segments']} segments")

        dl1, dl2 = st.columns(2)
        with dl1:
            st.download_button(
                "⬇️ Excel",
                key=f"dx_{idx}",
                data=style_excel(df, sheet_name=stats["filename"][:31].replace(".xlsx", ""), glossary_ws=glossary_ws),
                file_name=stats["filename"].replace(".xlsx", "_processed.xlsx"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with dl2:
            st.download_button(
                "⬇️ CSV",
                key=f"dc_{idx}",
                data=df.to_csv(index=False).encode("utf-8-sig"),
                file_name=stats["filename"].replace(".xlsx", "_processed.csv"),
                mime="text/csv",
                use_container_width=True,
            )

st.divider()
st.markdown("### 📦 Bulk Export")

e1, e2, e3 = st.columns(3)
with e1:
    st.download_button(
        "⬇️ Download All as ZIP",
        data=build_zip(results),
        file_name="pipeline_processed_all.zip",
        mime="application/zip",
        use_container_width=True,
    )
with e2:
    st.download_button(
        "⬇️ Merged Excel",
        data=style_excel(all_dfs, "All_Segments", glossary_ws=next((g for _, _, g in results if g is not None), None)),
        file_name="ALL_FILES_MERGED.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with e3:
    st.download_button(
        "⬇️ Merged CSV",
        data=all_dfs.to_csv(index=False).encode("utf-8-sig"),
        file_name="ALL_FILES_MERGED.csv",
        mime="text/csv",
        use_container_width=True,
    )

st.success(
    f"✅ {len(results)} file(s) processed · {total_final} segments ready for linguists · {total_units:,} source units total"
)
